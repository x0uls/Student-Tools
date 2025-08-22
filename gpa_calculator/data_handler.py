import os
import openpyxl

DATA_FILE = os.path.join(os.path.dirname(__file__), "data", "gpa_data.xlsx")


def init_data_file():
    if not os.path.exists(DATA_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Semesters"
        # Example header row
        ws.append(["Semester", "Subject", "Credit Hours", "Grade"])
        wb.save(DATA_FILE)


def load_data():
    wb = openpyxl.load_workbook(DATA_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return rows[1:]  # skip header


def save_data(data_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Semesters"
    ws.append(["Semester", "Subject", "Credit Hours", "Grade"])
    for row in data_rows:
        ws.append(row)
    wb.save(DATA_FILE)
