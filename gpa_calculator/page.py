import customtkinter as ctk
import openpyxl
import os
from .semester_detail_page import SemesterDetailPage, GRADE_POINTS
from .chart import GPAChartPage


class GPACalculatorPage(ctk.CTkFrame):
    """Main GPA Calculator page with semester management and CGPA calculation."""
    
    def __init__(self, parent):
        """Initialize the GPA calculator page with UI components."""
        super().__init__(parent)
        self.parent = parent
        self.configure(fg_color="transparent")
        
        # Create header with title and CGPA display
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.pack(fill="x", padx=20, pady=(20, 15))
        
        title_label = ctk.CTkLabel(
            header_frame, 
            text="GPA Calculator", 
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="#ffffff"
        )
        title_label.pack(pady=(0, 10))

        cgpa_frame = ctk.CTkFrame(header_frame, fg_color="#1a2b4c", corner_radius=10)
        cgpa_frame.pack(fill="x")
        
        self.cgpa_label = ctk.CTkLabel(
            cgpa_frame, 
            text="Total CGPA: 0.00", 
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#ffffff"
        )
        self.cgpa_label.pack(pady=12, padx=15)

        # Create action buttons
        button_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_frame.pack(fill="x", padx=20, pady=(0, 15))
        button_frame.grid_columnconfigure((0, 1), weight=1)

        self.add_sem_btn = ctk.CTkButton(
            button_frame,
            text="Add Semester",
            font=ctk.CTkFont(size=14),
            height=35,
            corner_radius=8,
            command=self.add_semester,
        )
        self.add_sem_btn.grid(row=0, column=0, padx=(0, 5), sticky="ew")

        self.chart_btn = ctk.CTkButton(
            button_frame,
            text="View Charts",
            font=ctk.CTkFont(size=14),
            height=35,
            corner_radius=8,
            command=self.open_chart_page,
        )
        self.chart_btn.grid(row=0, column=1, padx=(5, 0), sticky="ew")

        # Create scrollable frame for semester cards
        self.scrollable_frame = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scrollable_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        self.semesters = []  # List to store semester data
        self.load_from_excel()

    def add_semester(self):
        """Add a new semester with auto-generated name."""
        sem_number = len(self.semesters) + 1
        sem_name = f"Semester {sem_number}"

        sem_data = {
            "name": sem_name,
            "gpa": 0.0,
            "subjects": [],
            "detail_page": None,
            "card": None,
            "gpa_label": None,
        }
        self.semesters.append(sem_data)
        self._create_semester_card(sem_data)
        self._update_total_cgpa()
        self._update_chart()

    def _create_semester_card(self, sem):
        """Create a clickable semester card with name, GPA, and remove button."""
        card = ctk.CTkFrame(self.scrollable_frame, corner_radius=8, height=60)
        card.pack(fill="x", pady=3, padx=0)
        card.pack_propagate(False)
        sem["card"] = card

        content_frame = ctk.CTkFrame(card, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, padx=15, pady=10)

        left_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        left_frame.pack(side="left", fill="both", expand=True)

        name_label = ctk.CTkLabel(
            left_frame, 
            text=sem["name"], 
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#ffffff",
            anchor="w"
        )
        name_label.pack(fill="x", pady=(0, 3))
        sem["name_label"] = name_label

        gpa_label = ctk.CTkLabel(
            left_frame, 
            text=f"GPA: {sem['gpa']:.4f}", 
            font=ctk.CTkFont(size=12),
            text_color="#d1d5db",
            anchor="w"
        )
        gpa_label.pack(fill="x")
        sem["gpa_label"] = gpa_label

        remove_btn = ctk.CTkButton(
            content_frame,
            text="×",
            width=25,
            height=25,
            fg_color="#dc2626",
            hover_color="#b91c1c",
            corner_radius=12,
            border_width=0,
            text_color="white",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=lambda s=sem: self.remove_semester(s),
        )
        remove_btn.pack(side="right", padx=(10, 0))

        # Bind click to open semester
        def on_click(event, s=sem):
            self.open_semester(s)

        for widget in [card, content_frame, left_frame, name_label, gpa_label]:
            widget.bind("<Button-1>", on_click)

    def remove_semester(self, sem):
        """Remove semester and renumber remaining semesters."""
        if sem["card"]:
            sem["card"].destroy()
        if sem["detail_page"]:
            sem["detail_page"].destroy()
        
        self.semesters.remove(sem)
        self._update_total_cgpa()
        
        # Renumber remaining semesters
        for i, s in enumerate(self.semesters, start=1):
            s["name"] = f"Semester {i}"
            if s.get("name_label"):
                s["name_label"].configure(text=s["name"])
        
        self._update_chart()

    def open_semester(self, sem):
        """Open semester detail page for editing subjects."""
        if sem["detail_page"] is None:
            sem["detail_page"] = SemesterDetailPage(
                self.parent,
                sem["name"],
                lambda dp=None: self.close_semester(sem),
                main_page=self,
                existing_subjects=sem["subjects"],
            )
            sem["detail_page"].place(relwidth=1, relheight=1)
        else:
            sem["detail_page"].place(relwidth=1, relheight=1)

        sem["detail_page"].lift()

    def update_semester_subjects(self, semester_name, subjects):
        """Update semester data and recalculate GPA."""
        sem = next(s for s in self.semesters if s["name"] == semester_name)
        sem["subjects"] = subjects
        sem["gpa"] = self._calculate_semester_gpa(subjects)

        if sem["gpa_label"]:
            sem["gpa_label"].configure(text=f"GPA: {sem['gpa']:.4f}")

        self._update_total_cgpa()
        self._update_chart()

    def _calculate_semester_gpa(self, subjects):
        """Calculate GPA for a list of subjects."""
        if not subjects:
            return 0.0
        total_points = sum(GRADE_POINTS[s["grade"]] * s["credit"] for s in subjects)
        total_credits = sum(s["credit"] for s in subjects)
        return total_points / total_credits if total_credits else 0.0

    def _calculate_cgpa_tarumt(self, all_subjects):
        """Calculate CGPA using TARUMT rules: exclude failed subject credits only once."""
        if not all_subjects:
            return 0.0
        
        # Group subjects by name (case-insensitive)
        subject_groups = {}
        for subject in all_subjects:
            name = subject["name"].strip().lower()
            if name not in subject_groups:
                subject_groups[name] = []
            subject_groups[name].append(subject)
        
        total_points = 0.0
        total_credits = 0.0
        
        for subject_name, subjects in subject_groups.items():
            has_failed = any(s["grade"] == "F" for s in subjects)
            
            # Add all grade points (including F = 0.0)
            for subject in subjects:
                total_points += GRADE_POINTS[subject["grade"]] * subject["credit"]
            
            # Exclude failed subject credits only once
            if has_failed:
                for i, subject in enumerate(subjects):
                    if i == 0:  # Skip first attempt
                        continue
                    total_credits += subject["credit"]
            else:
                for subject in subjects:
                    total_credits += subject["credit"]
        
        return total_points / total_credits if total_credits > 0 else 0.0

    def _update_total_cgpa(self, save_data=True):
        """Calculate and display total CGPA using TARUMT rules."""
        all_subjects = [subj for sem in self.semesters for subj in sem["subjects"]]
        cgpa = self._calculate_cgpa_tarumt(all_subjects)
        self.cgpa_label.configure(text=f"Total CGPA: {cgpa:.4f}")
        if save_data:
            self.save_to_excel()

    def show_main_page(self):
        """Show the main GPA calculator page."""
        self.place(relwidth=1, relheight=1)

    def close_semester(self, sem):
        """Close semester detail page and return to main page."""
        if sem["detail_page"]:
            sem["detail_page"].place_forget()
        self.show_main_page()

    def open_chart_page(self):
        """Open or refresh the GPA trend chart page."""
        if not hasattr(self, "chart_page") or self.chart_page is None:
            self.chart_page = GPAChartPage(
                self.parent, self.semesters, go_back_callback=self.show_main_page
            )
            self.chart_page.place(relwidth=1, relheight=1)
        else:
            self.chart_page.draw_chart()
            self.chart_page.place(relwidth=1, relheight=1)

        self.chart_page.lift()

    def _update_chart(self):
        """Update chart data and redraw if chart page exists."""
        if hasattr(self, "chart_page") and self.chart_page is not None:
            self.chart_page.semesters = self.semesters
            self.chart_page.draw_chart()

    def save_to_excel(self, filename="gpa_data.xlsx"):
        """Save all semester and subject data to Excel file."""
        folder = os.path.dirname(__file__)
        os.makedirs(folder, exist_ok=True)
        file_path = os.path.join(folder, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GPA Data"
        ws.append(["Semester", "Subject", "Credit", "Grade", "GPA"])

        for sem in self.semesters:
            for subj in sem["subjects"]:
                ws.append([
                    sem["name"],
                    subj["name"],
                    subj["credit"],
                    subj["grade"],
                    f"{sem['gpa']:.4f}",
                ])

        wb.save(file_path)
        print(f"Data saved automatically to {os.path.abspath(file_path)}")

    def load_from_excel(self, filename="gpa_data.xlsx"):
        """Load semester and subject data from Excel file."""
        folder = os.path.dirname(__file__)
        file_path = os.path.join(folder, filename)
        if not os.path.exists(file_path):
            return

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        data_rows = list(ws.iter_rows(values_only=True))[1:]  # Skip header

        semesters_dict = {}
        for sem_name, subj_name, credit, grade, gpa in data_rows:
            if sem_name not in semesters_dict:
                semesters_dict[sem_name] = {
                    "name": sem_name,
                    "gpa": 0.0,
                    "subjects": [],
                    "detail_page": None,
                    "card": None,
                    "gpa_label": None,
                }
            semesters_dict[sem_name]["subjects"].append({
                "name": subj_name,
                "credit": float(credit),
                "grade": grade,
            })

        self.semesters.clear()
        for sem in semesters_dict.values():
            sem["gpa"] = self._calculate_semester_gpa(sem["subjects"])
            self.semesters.append(sem)
            self._create_semester_card(sem)

        self._update_total_cgpa(save_data=False)
        print(f"Data loaded from {os.path.abspath(file_path)}")
