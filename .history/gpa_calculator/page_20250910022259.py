import customtkinter as ctk
import openpyxl
import os
from .semester_detail_page import SemesterDetailPage, GRADE_POINTS
from .chart import GPAChartPage


class GPACalculatorPage(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent

        # Configure main frame
        self.configure(fg_color="transparent")
        
        # Header section - clean and minimal
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.pack(fill="x", pady=(30, 20), padx=30)
        
        # Main title - minimalist typography with container
        title_container = ctk.CTkFrame(header_frame, fg_color="#fef3c7", corner_radius=10)
        title_container.pack(pady=(0, 15))
        
        title_label = ctk.CTkLabel(
            title_container, 
            text="GPA Calculator", 
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color="#92400e"
        )
        title_label.pack(pady=12, padx=20)

        # CGPA display - prominent but clean
        cgpa_container = ctk.CTkFrame(header_frame, fg_color="#fef3c7", corner_radius=12)
        cgpa_container.pack(pady=(0, 10))
        
        self.cgpa_label = ctk.CTkLabel(
            cgpa_container, 
            text="Total CGPA: 0.00", 
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="#92400e"
        )
        self.cgpa_label.pack(pady=15, padx=20)

        # Semester cards section - clean scrollable area
        self.scrollable_frame = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scrollable_frame.pack(fill="both", expand=True, padx=30, pady=(0, 20))

        # Inner frame for the semester cards
        self.semester_frame = ctk.CTkFrame(self.scrollable_frame, fg_color="transparent")
        self.semester_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Keep track of semesters
        self.semesters = []

        # Bottom section - minimal action area
        bottom_frame = ctk.CTkFrame(self, fg_color="transparent")
        bottom_frame.pack(fill="x", pady=(0, 30), padx=30)

        # Add Semester Button - clean and prominent
        self.add_sem_btn = ctk.CTkButton(
            bottom_frame,
            text="Add Semester",
            font=ctk.CTkFont(size=15, weight="bold"),
            height=36,
            width=180,
            corner_radius=18,
            fg_color="#6366f1",
            hover_color="#4f46e5",
            text_color="white",
            command=self.add_semester,
        )
        self.add_sem_btn.pack(pady=(0, 15))

        # Chart link - subtle and clean
        self.chart_link = ctk.CTkLabel(
            bottom_frame,
            text="View GPA Chart",
            font=ctk.CTkFont(size=13),
            text_color="#6b7280",
            cursor="hand2",
        )
        self.chart_link.pack()

        self.chart_link.bind("<Button-1>", lambda e: self.open_chart_page())

        self.load_from_excel()

    def add_semester(self):
        sem_number = len(self.semesters) + 1
        sem_name = f"Semester {sem_number}"

        sem_data = {
            "name": sem_name,
            "gpa": 0.0,
            "subjects": [],
            "detail_page": None,
            "card": None,
            "gpa_label": None,
        }
        self.semesters.append(sem_data)
        self._create_semester_card(sem_data)
        self._update_total_cgpa()

    def _create_semester_card(self, sem):
        # Clean, minimal card design
        card = ctk.CTkFrame(self.semester_frame, corner_radius=12, height=70)
        card.pack(fill="x", pady=6, padx=0)
        card.pack_propagate(False)
        sem["card"] = card

        # Main content container
        content_frame = ctk.CTkFrame(card, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, padx=20, pady=12)

        # Left side: semester info
        info_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        info_frame.pack(side="left", fill="both", expand=True)

        # Semester name - clean typography
        name_label = ctk.CTkLabel(
            info_frame, 
            text=sem["name"], 
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#1f2937",
            anchor="w"
        )
        name_label.pack(fill="x", pady=(0, 3))
        sem["name_label"] = name_label

        # GPA display - subtle styling
        gpa_label = ctk.CTkLabel(
            info_frame, 
            text=f"GPA: {sem['gpa']:.2f}", 
            font=ctk.CTkFont(size=13),
            text_color="#6b7280",
            anchor="w"
        )
        gpa_label.pack(fill="x")
        sem["gpa_label"] = gpa_label

        # Right side: remove button - minimal and clean
        remove_btn = ctk.CTkButton(
            content_frame,
            text="×",
            width=24,
            height=24,
            fg_color="#f87171",
            hover_color="#ef4444",
            corner_radius=12,
            border_width=0,
            text_color="white",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=lambda s=sem: self.remove_semester(s),
        )
        remove_btn.pack(side="right", padx=(10, 0))

        # Bind click to open semester
        def on_click(event, s=sem):
            self.open_semester(s)

        for widget in [card, content_frame, info_frame, name_label, gpa_label]:
            widget.bind("<Button-1>", on_click)

    def remove_semester(self, sem):
        """Remove semester card and data."""
        # Destroy the card
        if sem["card"]:
            sem["card"].destroy()
        # Destroy detail page if it exists
        if sem["detail_page"]:
            sem["detail_page"].destroy()
        # Remove from list
        self.semesters.remove(sem)
        # Update total CGPA
        self._update_total_cgpa()
        # Re-number remaining semesters
        for i, s in enumerate(self.semesters, start=1):
            s["name"] = f"Semester {i}"
            if s.get("name_label"):
                s["name_label"].configure(text=s["name"])

    def open_semester(self, sem):
        if sem["detail_page"] is None:
            sem["detail_page"] = SemesterDetailPage(
                self.parent,
                sem["name"],
                lambda dp=None: self.close_semester(sem),
                main_page=self,
                existing_subjects=sem["subjects"],
            )
            sem["detail_page"].place(relwidth=1, relheight=1)
        else:
            sem["detail_page"].place(relwidth=1, relheight=1)

        sem["detail_page"].lift()

    def update_semester_subjects(self, semester_name, subjects):
        sem = next(s for s in self.semesters if s["name"] == semester_name)
        sem["subjects"] = subjects

        # Update GPA
        total_points = sum(GRADE_POINTS[s["grade"]] * s["credit"] for s in subjects)
        total_credits = sum(s["credit"] for s in subjects)
        sem["gpa"] = total_points / total_credits if total_credits else 0.0

        # Update GPA label
        if sem["gpa_label"]:
            sem["gpa_label"].configure(text=f"GPA: {sem['gpa']:.2f}")

        self._update_total_cgpa()
        self.save_to_excel()

    def _update_total_cgpa(self):
        """Calculate CGPA from all semesters."""
        total_points = 0
        total_credits = 0
        for sem in self.semesters:
            for subj in sem["subjects"]:
                total_points += GRADE_POINTS[subj["grade"]] * subj["credit"]
                total_credits += subj["credit"]
        cgpa = total_points / total_credits if total_credits else 0.0
        self.cgpa_label.configure(text=f"Total CGPA: {cgpa:.2f}")
        self.save_to_excel()

    def show_main_page(self):
        self.place(relwidth=1, relheight=1)

    def close_semester(self, sem):
        if sem["detail_page"]:
            sem["detail_page"].place_forget()
        self.show_main_page()

    def open_chart_page(self):
        if not hasattr(self, "chart_page") or self.chart_page is None:
            self.chart_page = GPAChartPage(
                self.parent, self.semesters, go_back_callback=self.show_main_page
            )
            self.chart_page.place(relwidth=1, relheight=1)
        else:
            self.chart_page.draw_chart()
            self.chart_page.place(relwidth=1, relheight=1)

        self.chart_page.lift()

    def save_to_excel(self, filename="gpa_data.xlsx"):
        """Save all semester/subject data into an Excel file."""
        folder = os.path.join(os.path.dirname(__file__))
        os.makedirs(folder, exist_ok=True)  # create folder if it doesn't exist
        file_path = os.path.join(folder, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GPA Data"

        # Header row
        ws.append(["Semester", "Subject", "Credit", "Grade", "GPA"])

        for sem in self.semesters:
            for subj in sem["subjects"]:
                ws.append(
                    [
                        sem["name"],
                        subj["name"],
                        subj["credit"],
                        subj["grade"],
                        f"{sem['gpa']:.2f}",
                    ]
                )

        # Save workbook
        wb.save(file_path)
        print(f"Data saved automatically to {os.path.abspath(file_path)}")

    def load_from_excel(self, filename="gpa_data.xlsx"):
        folder = os.path.dirname(__file__)
        file_path = os.path.join(folder, filename)
        if not os.path.exists(file_path):
            return

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        data_rows = list(ws.iter_rows(values_only=True))[1:]  # skip header

        semesters_dict = {}
        for sem_name, subj_name, credit, grade, gpa in data_rows:
            if sem_name not in semesters_dict:
                semesters_dict[sem_name] = {
                    "name": sem_name,
                    "gpa": 0.0,
                    "subjects": [],
                    "detail_page": None,
                    "card": None,
                    "gpa_label": None,
                }
            semesters_dict[sem_name]["subjects"].append(
                {
                    "name": subj_name,
                    "credit": float(credit),
                    "grade": grade,
                }
            )

        # Clear and rebuild
        self.semesters.clear()
        for sem in semesters_dict.values():
            total_points = sum(
                GRADE_POINTS[s["grade"]] * s["credit"] for s in sem["subjects"]
            )
            total_credits = sum(s["credit"] for s in sem["subjects"])
            sem["gpa"] = total_points / total_credits if total_credits else 0.0

            self.semesters.append(sem)
            self._create_semester_card(sem)

        self._update_total_cgpa()
        print(f"Data loaded from {os.path.abspath(file_path)}")
